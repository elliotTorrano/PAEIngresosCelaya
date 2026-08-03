from pathlib import Path

import openpyxl

from app.auth.crypto_certs import generate_certificate_bundle, load_bundle
from app.config import AUTH_TYPE_CERTIFICADO, AUTH_TYPE_PASSWORD, ROLE_ABOGADO, ROLE_AGENTE_PAE
from app.db.repositories import users as users_repo
from app.excel_io.requerimientos_export import export_for_abogado
from app.excel_io.requerimientos_import import parse_agente_export_file, parse_requerimientos_file


def _write_raw_excel(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["irrelevant title row"])  # fila 1 (se omite)
    ws.append(["A", "FOLIO", "CTA PREDIAL", "CONTRIBUYENTE", "E", "DOMICILIO"])  # fila 2 (se omite)
    ws.append(["x1", "F-001", "CP-001", "Juan Pérez", "y1", "Calle 1"])
    ws.append(["x2", "F-002", "CP-002", "María López", "y2", "Calle 2"])
    ws.append(["TOTAL", "", "", "", "", ""])  # última fila (se omite)
    wb.save(path)


def test_parse_requerimientos_file_skips_rows_and_maps_columns(tmp_path):
    path = tmp_path / "origen.xlsx"
    _write_raw_excel(path)

    result = parse_requerimientos_file(path)

    assert result.row_count == 2
    assert result.rows[0] == {
        "folio": "F-001",
        "cta_predial": "CP-001",
        "contribuyente": "Juan Pérez",
        "domicilio": "Calle 1",
    }
    assert result.rows[1]["folio"] == "F-002"


def test_parse_requerimientos_file_folio_numerico_sin_decimales(tmp_path):
    """Cuando Excel guarda el FOLIO como número (no como texto), openpyxl lo
    entrega como float (p. ej. 1234.0) -- no debe verse "1234.0" en el resultado."""
    path = tmp_path / "origen_numerico.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["irrelevant title row"])
    ws.append(["A", "FOLIO", "CTA PREDIAL", "CONTRIBUYENTE", "E", "DOMICILIO"])
    ws.append(["x1", 1234, "CP-001", "Juan Pérez", "y1", "Calle 1"])
    ws.append(["TOTAL", "", "", "", "", ""])
    wb.save(path)

    result = parse_requerimientos_file(path)

    assert result.rows[0]["folio"] == "1234"


def test_parse_requerimientos_file_supports_legacy_xls(tmp_path):
    import xlwt

    path = tmp_path / "origen.xls"
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Hoja1")
    rows = [
        ["irrelevant title row"],
        ["A", "FOLIO", "CTA PREDIAL", "CONTRIBUYENTE", "E", "DOMICILIO"],
        ["x1", "F-001", "CP-001", "Juan Pérez", "y1", "Calle 1"],
        ["x2", "F-002", "CP-002", "María López", "y2", "Calle 2"],
        ["TOTAL", "", "", "", "", ""],
    ]
    for r, row in enumerate(rows):
        for c, value in enumerate(row):
            ws.write(r, c, value)
    wb.save(str(path))

    result = parse_requerimientos_file(path)

    assert result.row_count == 2
    assert result.rows[0] == {
        "folio": "F-001",
        "cta_predial": "CP-001",
        "contribuyente": "Juan Pérez",
        "domicilio": "Calle 1",
    }
    assert result.rows[1]["folio"] == "F-002"


def test_parse_agente_export_file_reads_clean_headers(db, tmp_path):
    agente = users_repo.create_user(
        username="agente1", role=ROLE_AGENTE_PAE, full_name="Agente Uno", email="a@a.com",
        auth_type=AUTH_TYPE_CERTIFICADO,
    )
    pfx_bytes, cert_public_pem, cert_serial = generate_certificate_bundle(
        username=agente.username, full_name=agente.full_name, password="clave-segura"
    )
    users_repo.set_certificate(agente.id, cert_public_pem=cert_public_pem, cert_serial=cert_serial)
    agente = users_repo.get_by_id(agente.id)
    private_key, _certificate = load_bundle(pfx_bytes, "clave-segura")

    abogado = users_repo.create_user(
        username="abogado1", role=ROLE_ABOGADO, full_name="Abogado Uno", email="b@b.com",
        auth_type=AUTH_TYPE_PASSWORD, password_hash="x", password_salt="y",
    )

    path = tmp_path / "exportado.mcdiep"
    export_for_abogado(
        [{"folio": "F-001", "cta_predial": "CP-001", "contribuyente": "Juan Pérez", "domicilio": "Calle 1"}],
        path, agente=agente, abogado=abogado, private_key=private_key,
    )

    result = parse_agente_export_file(path, abogado=abogado)

    assert result.rows == [
        {"folio": "F-001", "cta_predial": "CP-001", "contribuyente": "Juan Pérez", "domicilio": "Calle 1"}
    ]
    assert result.agente.username == "agente1"
    assert result.document_uuid
    assert result.file_hash
