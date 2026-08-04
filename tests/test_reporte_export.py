import openpyxl

from app.db.repositories import reporte_mandamientos as reporte_mand_repo
from app.db.repositories import reporte_requerimientos as reporte_req_repo
from app.excel_io import reporte_mandamientos_export, reporte_requerimientos_export


def test_export_reporte_xlsx_requerimientos_writes_19_columns_in_order(db, tmp_path):
    reporte_req_repo.add_source_rows(
        [{"folio": "F-001", "cta_predial": "CP-001", "contribuyente": "Juan Pérez", "domicilio": "Calle 1", "adeudo": "1500.00"}],
        lista_numero="LISTA-1", fecha_impreso="01/01/2026", source_filename="origen.xlsx",
    )
    reporte_req_repo.add_revision_rows(
        [{
            "folio": "F-001", "despacho": "Despacho Uno", "fecha_citatorio": "02/01/2026",
            "recibe_citatorio": "EN PUERTA", "fecha_notificacion": "03/01/2026", "quien_recibe": "NOMBRE",
            "observaciones": "Sin novedad.",
        }]
    )
    row_id = reporte_req_repo.list_rows()[0].id
    reporte_req_repo.update_manual_field(row_id, "observaciones_area", "Revisado.")
    reporte_req_repo.update_manual_field(row_id, "motivo_suspension", "N/A")
    reporte_req_repo.update_manual_field(row_id, "fecha_extrajudicial", "10/01/2026")
    reporte_req_repo.update_manual_field(row_id, "domicilio_notificacion", "Calle 1 Bis")
    reporte_req_repo.bulk_set_fecha_entrega("LISTA-1", "05/01/2026")
    reporte_req_repo.update_manual_field(row_id, "fecha_recepcion", "06/01/2026")

    output_path = tmp_path / "reporte.xlsx"
    reporte_requerimientos_export.export_reporte_xlsx(reporte_req_repo.list_rows(), output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.active
    header_row = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header_row == reporte_requerimientos_export.HEADERS
    assert len(header_row) == 19

    data_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    assert data_row == (
        "LISTA-1", "F-001", "CP-001", "Juan Pérez", "Calle 1", "Calle 1 Bis", "1500.00",
        "Despacho Uno", "01/01/2026", "05/01/2026", "06/01/2026", "02/01/2026", "EN PUERTA",
        "03/01/2026", "NOMBRE", "Sin novedad.", "Revisado.", "10/01/2026", "N/A",
    )


def test_export_reporte_xlsx_mandamientos_writes_17_columns_no_domicilio(db, tmp_path):
    reporte_mand_repo.add_source_rows(
        [{"folio": "F-002", "cta_predial": "CP-002", "contribuyente": "María López", "adeudo": "900.00"}],
        lista_numero="LISTA-2", fecha_impreso="01/01/2026", source_filename="origen.xlsx",
    )

    output_path = tmp_path / "reporte_mandamientos.xlsx"
    reporte_mandamientos_export.export_reporte_xlsx(reporte_mand_repo.list_rows(), output_path)

    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.active
    header_row = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header_row == reporte_mandamientos_export.HEADERS
    assert len(header_row) == 17
    assert "DOMICILIO DE UBICACIÓN" not in header_row

    data_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    assert data_row[0:5] == ("LISTA-2", "F-002", "CP-002", "María López", "900.00")
