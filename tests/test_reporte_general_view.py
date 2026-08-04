from unittest.mock import MagicMock, patch

from PySide6.QtWidgets import QDialog

from app.db.repositories import reporte_requerimientos as reporte_repo
from app.db.repositories import settings as settings_repo
from app.ui.reporteador.reporte_general_view import ReporteGeneralView
from app.ui.reporteador.reporte_requerimientos_view import (
    COL_FOLIO,
    COL_OBS_AREA,
    ReporteRequerimientosView,
)


def _write_source_file(path, folio="F-001", adeudo="1500.00"):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Título"])
    ws.append(["A", "FOLIO", "CTA PREDIAL", "CONTRIBUYENTE", "E", "DOMICILIO", "G", "H", "I", "J", "ADEUDO"])
    ws.append(["x1", folio, "CP-001", "Juan Pérez", "y1", "Calle 1", "", "", "", "", adeudo])
    ws.append(["TOTAL", "", "", "", "", "", "", "", "", "", ""])
    wb.save(path)


def _mock_dialog(accepted: bool, **attrs):
    dialog = MagicMock()
    dialog.exec.return_value = QDialog.DialogCode.Accepted if accepted else QDialog.DialogCode.Rejected
    dialog.DialogCode = QDialog.DialogCode
    for key, value in attrs.items():
        setattr(dialog, key, value)
    return dialog


def test_import_source_creates_rows_via_assign_lista_dialog(qapp, db, tmp_path):
    source_path = tmp_path / "origen.xlsx"
    _write_source_file(source_path)

    view = ReporteRequerimientosView(user=None)
    dialog = _mock_dialog(True, result_by_filename={"origen.xlsx": ("LISTA-1", "01/01/2026")})

    with patch(
        "app.ui.reporteador.reporte_requerimientos_view.QFileDialog.getOpenFileNames",
        return_value=([str(source_path)], ""),
    ), patch(
        "app.ui.reporteador.reporte_requerimientos_view.AssignListaDialog", return_value=dialog,
    ), patch("app.ui.reporteador.reporte_requerimientos_view.QMessageBox.information"):
        view._on_import_source()

    rows = reporte_repo.list_rows()
    assert len(rows) == 1
    assert rows[0].folio == "F-001"
    assert rows[0].lista_numero == "LISTA-1"
    assert rows[0].adeudo == "1500.00"
    assert view.table.rowCount() == 1


def test_import_source_cancelled_dialog_does_not_persist(qapp, db, tmp_path):
    source_path = tmp_path / "origen.xlsx"
    _write_source_file(source_path)

    view = ReporteRequerimientosView(user=None)
    dialog = _mock_dialog(False)

    with patch(
        "app.ui.reporteador.reporte_requerimientos_view.QFileDialog.getOpenFileNames",
        return_value=([str(source_path)], ""),
    ), patch(
        "app.ui.reporteador.reporte_requerimientos_view.AssignListaDialog", return_value=dialog,
    ):
        view._on_import_source()

    assert reporte_repo.list_rows() == []


def test_import_source_duplicate_folio_reported_and_not_overwritten(qapp, db, tmp_path):
    reporte_repo.add_source_rows(
        [{"folio": "F-001", "cta_predial": "CP-001", "contribuyente": "Juan Pérez", "domicilio": "Calle 1", "adeudo": "1500.00"}],
        lista_numero="LISTA-1", fecha_impreso="01/01/2026", source_filename="origen.xlsx",
    )
    source_path = tmp_path / "origen2.xlsx"
    _write_source_file(source_path, adeudo="9999.00")

    view = ReporteRequerimientosView(user=None)
    dialog = _mock_dialog(True, result_by_filename={"origen2.xlsx": ("LISTA-2", "05/01/2026")})

    with patch(
        "app.ui.reporteador.reporte_requerimientos_view.QFileDialog.getOpenFileNames",
        return_value=([str(source_path)], ""),
    ), patch(
        "app.ui.reporteador.reporte_requerimientos_view.AssignListaDialog", return_value=dialog,
    ), patch("app.ui.reporteador.reporte_requerimientos_view.QMessageBox.information") as mock_info:
        view._on_import_source()

    mock_info.assert_called_once()
    message = mock_info.call_args.args[2]
    assert "F-001" in message
    assert reporte_repo.list_rows()[0].adeudo == "1500.00"


def test_manual_field_edit_persists_and_syncs_master_file(qapp, db, tmp_path):
    reporte_repo.add_source_rows(
        [{"folio": "F-001", "cta_predial": "CP-001", "contribuyente": "Juan Pérez", "domicilio": "Calle 1", "adeudo": "1500.00"}],
        lista_numero="LISTA-1", fecha_impreso="01/01/2026", source_filename="origen.xlsx",
    )
    master_path = tmp_path / "maestro.xlsx"
    settings_repo.set(settings_repo.KEY_REPORTE_REQUERIMIENTOS_EXCEL_PATH, str(master_path))

    view = ReporteRequerimientosView(user=None)
    row_id = view._rows[0].id
    edit = view.table.cellWidget(0, COL_OBS_AREA)
    edit.setText("Revisado por control interno.")
    edit.editingFinished.emit()

    persisted = reporte_repo.list_rows()[0]
    assert persisted.observaciones_area == "Revisado por control interno."
    assert master_path.exists()

    import openpyxl
    wb = openpyxl.load_workbook(master_path)
    values = next(wb.active.iter_rows(min_row=2, max_row=2, values_only=True))
    assert "Revisado por control interno." in values


def test_asignar_fecha_entrega_updates_matching_rows(qapp, db):
    reporte_repo.add_source_rows(
        [{"folio": "F-001", "cta_predial": "CP-001", "contribuyente": "Juan Pérez", "domicilio": "Calle 1", "adeudo": "1500.00"}],
        lista_numero="LISTA-1", fecha_impreso="01/01/2026", source_filename="origen.xlsx",
    )
    view = ReporteRequerimientosView(user=None)
    dialog = _mock_dialog(True, selected_lista="LISTA-1", selected_fecha="20/01/2026")

    with patch(
        "app.ui.reporteador.reporte_requerimientos_view.AsignarFechaEntregaDialog", return_value=dialog,
    ), patch("app.ui.reporteador.reporte_requerimientos_view.QMessageBox.information"):
        view._on_asignar_fecha_entrega()

    assert reporte_repo.list_rows()[0].fecha_entrega == "20/01/2026"


def test_asignar_fecha_entrega_warns_when_no_lista(qapp, db):
    view = ReporteRequerimientosView(user=None)

    with patch("app.ui.reporteador.reporte_requerimientos_view.QMessageBox.information") as mock_info:
        view._on_asignar_fecha_entrega()

    mock_info.assert_called_once()


def test_search_by_folio_selects_matching_row(qapp, db):
    reporte_repo.add_source_rows(
        [
            {"folio": "F-001", "cta_predial": "CP-001", "contribuyente": "Juan Pérez", "domicilio": "Calle 1", "adeudo": "1"},
            {"folio": "F-002", "cta_predial": "CP-002", "contribuyente": "María López", "domicilio": "Calle 2", "adeudo": "2"},
        ],
        lista_numero="LISTA-1", fecha_impreso="01/01/2026", source_filename="origen.xlsx",
    )
    view = ReporteRequerimientosView(user=None)

    view.search_input.setText("f-002")
    view._on_search()

    assert view.table.currentRow() == 1


def test_export_copy_writes_file_without_changing_master_setting(qapp, db, tmp_path):
    reporte_repo.add_source_rows(
        [{"folio": "F-001", "cta_predial": "CP-001", "contribuyente": "Juan Pérez", "domicilio": "Calle 1", "adeudo": "1"}],
        lista_numero="LISTA-1", fecha_impreso="01/01/2026", source_filename="origen.xlsx",
    )
    view = ReporteRequerimientosView(user=None)
    copy_path = tmp_path / "copia.xlsx"

    with patch(
        "app.ui.reporteador.reporte_requerimientos_view.QFileDialog.getSaveFileName",
        return_value=(str(copy_path), ""),
    ), patch("app.ui.reporteador.reporte_requerimientos_view.QMessageBox.information"):
        view._on_export_copy()

    assert copy_path.exists()
    assert settings_repo.get(settings_repo.KEY_REPORTE_REQUERIMIENTOS_EXCEL_PATH) is None


def test_reporte_general_view_has_both_tabs(qapp, db):
    view = ReporteGeneralView(user=None)
    assert view.tab_requerimientos is not None
    assert view.tab_mandamientos is not None
