from unittest.mock import MagicMock, patch

import openpyxl
from PySide6.QtWidgets import QDialog, QMessageBox

from app.auth.crypto_certs import generate_certificate_bundle, load_bundle
from app.config import AUTH_TYPE_CERTIFICADO, AUTH_TYPE_PASSWORD, ROLE_ABOGADO, ROLE_AGENTE_PAE
from app.db.repositories import users as users_repo
from app.ui.agente.mandamientos_generar_view import MandamientosGenerarView
from app.ui.agente.requerimientos_generar_view import RequerimientosGenerarView
from app.ui.widgets.certificate_confirm_dialog import CertificateConfirmDialog


def _write_valid_requerimientos_file(path, folio="F-001"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Título"])
    ws.append(["A", "FOLIO", "CTA PREDIAL", "CONTRIBUYENTE", "E", "DOMICILIO"])
    ws.append(["x1", folio, "CP-001", "Juan Pérez", "y1", "Calle 1"])
    ws.append(["TOTAL", "", "", "", "", ""])
    wb.save(path)


def _write_valid_mandamientos_file(path, folio="F-001"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Título"])
    ws.append(["A", "FOLIO", "CTA PREDIAL", "CONTRIBUYENTE"])
    ws.append(["x1", folio, "CP-001", "Juan Pérez"])
    ws.append(["TOTAL", "", "", ""])
    wb.save(path)


def _make_agente_with_pfx(tmp_path, password="clave-agente"):
    agente = users_repo.create_user(
        username="agente1", role=ROLE_AGENTE_PAE, full_name="Agente Uno", email="a@a.com",
        auth_type=AUTH_TYPE_CERTIFICADO,
    )
    pfx_bytes, cert_public_pem, cert_serial = generate_certificate_bundle(
        username=agente.username, full_name=agente.full_name, password=password
    )
    users_repo.set_certificate(agente.id, cert_public_pem=cert_public_pem, cert_serial=cert_serial)
    agente = users_repo.get_by_id(agente.id)
    agente_pfx_path = tmp_path / "agente1.pfx"
    agente_pfx_path.write_bytes(pfx_bytes)

    users_repo.create_user(
        username="abogado1", role=ROLE_ABOGADO, full_name="Abogado Uno", email="b@b.com",
        auth_type=AUTH_TYPE_PASSWORD, password_hash="x", password_salt="y",
    )
    return agente, agente_pfx_path


def _make_confirm_exec(pfx_path, password):
    def _exec(self):
        self._cert_path = str(pfx_path)
        self.password_input.setText(password)
        self._on_confirm()
        return self.result()

    return _exec


def test_requerimientos_export_writes_backup_xlsx(qapp, db, tmp_path):
    agente, agente_pfx_path = _make_agente_with_pfx(tmp_path)
    source_path = tmp_path / "lote.xlsx"
    _write_valid_requerimientos_file(source_path)
    dest_folder = tmp_path / "destino"
    dest_folder.mkdir()

    view = RequerimientosGenerarView(agente)
    with patch(
        "app.ui.agente.requerimientos_generar_view.QFileDialog.getOpenFileNames",
        return_value=([str(source_path)], ""),
    ):
        view._on_select_files()

    with patch.object(
        CertificateConfirmDialog, "exec", _make_confirm_exec(agente_pfx_path, "clave-agente")
    ), patch(
        "app.ui.agente.requerimientos_generar_view.QFileDialog.getExistingDirectory",
        return_value=str(dest_folder),
    ), patch(
        "app.ui.agente.requerimientos_generar_view.QMessageBox.question",
        return_value=QMessageBox.StandardButton.Yes,
    ), patch("app.ui.agente.requerimientos_generar_view.QMessageBox.information"):
        view._on_export()

    xlsx_matches = list(dest_folder.glob("*.xlsx"))
    assert len(xlsx_matches) == 1
    assert len(list(dest_folder.glob("*.mcdiep"))) == 1
    assert len(list(dest_folder.glob("*.pdf"))) == 1

    wb = openpyxl.load_workbook(xlsx_matches[0])
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("FOLIO", "CTA PREDIAL", "CONTRIBUYENTE", "DOMICILIO")
    assert rows[1] == ("F-001", "CP-001", "Juan Pérez", "Calle 1")


def test_mandamientos_export_writes_backup_xlsx(qapp, db, tmp_path):
    agente, agente_pfx_path = _make_agente_with_pfx(tmp_path)
    source_path = tmp_path / "lote.xlsx"
    _write_valid_mandamientos_file(source_path)
    dest_folder = tmp_path / "destino"
    dest_folder.mkdir()

    view = MandamientosGenerarView(agente)
    with patch(
        "app.ui.agente.mandamientos_generar_view.QFileDialog.getOpenFileNames",
        return_value=([str(source_path)], ""),
    ):
        view._on_select_files()

    with patch.object(
        CertificateConfirmDialog, "exec", _make_confirm_exec(agente_pfx_path, "clave-agente")
    ), patch(
        "app.ui.agente.mandamientos_generar_view.QFileDialog.getExistingDirectory",
        return_value=str(dest_folder),
    ), patch(
        "app.ui.agente.mandamientos_generar_view.QMessageBox.question",
        return_value=QMessageBox.StandardButton.Yes,
    ), patch("app.ui.agente.mandamientos_generar_view.QMessageBox.information"):
        view._on_export()

    xlsx_matches = list(dest_folder.glob("*.xlsx"))
    assert len(xlsx_matches) == 1
    assert len(list(dest_folder.glob("*.mcdiep"))) == 1
    assert len(list(dest_folder.glob("*.pdf"))) == 1

    wb = openpyxl.load_workbook(xlsx_matches[0])
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("FOLIO", "CTA PREDIAL", "CONTRIBUYENTE")
    assert rows[1] == ("F-001", "CP-001", "Juan Pérez")


def test_dummy_export_does_not_write_backup_xlsx(qapp, db, tmp_path):
    agente = users_repo.create_user(
        username="agente_dummy", role=ROLE_AGENTE_PAE, full_name="Agente del PAE (prueba)", email=None,
        auth_type=AUTH_TYPE_PASSWORD, password_hash="x", password_salt="y",
    )
    users_repo.create_user(
        username="abogado1", role=ROLE_ABOGADO, full_name="Abogado Uno", email="b@b.com",
        auth_type=AUTH_TYPE_PASSWORD, password_hash="x", password_salt="y",
    )
    source_path = tmp_path / "lote.xlsx"
    _write_valid_requerimientos_file(source_path)
    dest_folder = tmp_path / "destino"
    dest_folder.mkdir()

    view = RequerimientosGenerarView(agente, dummy=True)
    with patch(
        "app.ui.agente.requerimientos_generar_view.QFileDialog.getOpenFileNames",
        return_value=([str(source_path)], ""),
    ):
        view._on_select_files()

    with patch(
        "app.ui.agente.requerimientos_generar_view.QFileDialog.getExistingDirectory",
        return_value=str(dest_folder),
    ), patch(
        "app.ui.agente.requerimientos_generar_view.QMessageBox.question",
        return_value=QMessageBox.StandardButton.Yes,
    ), patch("app.ui.agente.requerimientos_generar_view.QMessageBox.information"):
        view._on_export()

    assert list(dest_folder.glob("*.xlsx")) == []
