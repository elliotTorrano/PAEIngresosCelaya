from unittest.mock import patch

from app.config import AUTH_TYPE_CERTIFICADO, AUTH_TYPE_PASSWORD, ROLE_ABOGADO, ROLE_AGENTE_PAE
from app.db.repositories import revisiones as revisiones_repo
from app.db.repositories import users as users_repo
from app.db.repositories.requerimientos import RequerimientoRow
from app.excel_io.requerimientos_export import HEADERS_ABOGADO, export_captured
from app.ui.agente.requerimientos_revision_view import RequerimientosRevisionView


def _make_agente_abogado():
    agente = users_repo.create_user(
        username="agente1", role=ROLE_AGENTE_PAE, full_name="Agente Uno", email="a@a.com",
        auth_type=AUTH_TYPE_CERTIFICADO,
    )
    users_repo.create_user(
        username="abogado1", role=ROLE_ABOGADO, full_name="Abogado Uno", email="b@b.com",
        auth_type=AUTH_TYPE_PASSWORD, password_hash="x", password_salt="y",
    )
    return agente


def _write_captura_file(path, folio="F-001"):
    row = RequerimientoRow(
        id=1, batch_id=1, folio=folio, cta_predial="CP-001", contribuyente="Juan Pérez", domicilio="Calle 1",
        fecha_citatorio="01/01/2026", recibe_citatorio="EN PUERTA", recibe_citatorio_nombre=None,
        fecha_notificacion="02/01/2026", quien_recibe="EN PUERTA", quien_recibe_nombre=None,
        observaciones=None,
    )
    export_captured([row], path)


def _row(folio):
    return {
        "folio": folio, "cta_predial": None, "contribuyente": None, "domicilio": None,
        "fecha_citatorio": None, "recibe_citatorio": None, "recibe_citatorio_nombre": None,
        "fecha_notificacion": None, "quien_recibe": None, "quien_recibe_nombre": None,
    }


def _import_rows(agente_id, *, source_filename, rows):
    """Igual que hace la vista al importar: crea el `revision_import` y luego
    sus filas -- usado por las pruebas que necesitan datos preexistentes sin
    pasar por un diálogo de archivo real."""
    revision_import_id = revisiones_repo.create_revision_import(
        agente_id=agente_id, source_filename=source_filename, abogado_nombre=None, abogado_id=None,
    )
    revisiones_repo.add_revision_rows(
        agente_id=agente_id, revision_import_id=revision_import_id,
        source_filename=source_filename, abogado_nombre=None, abogado_id=None, rows=rows,
    )
    return revision_import_id


def test_import_revision_persists_rows_and_refreshes_table(qapp, db, tmp_path):
    agente = _make_agente_abogado()
    path = tmp_path / "captura.mcdiep"
    _write_captura_file(path)

    view = RequerimientosRevisionView(agente)
    with patch(
        "app.ui.agente.requerimientos_revision_view.QFileDialog.getOpenFileName",
        return_value=(str(path), ""),
    ), patch("app.ui.agente.requerimientos_revision_view.QMessageBox.information"):
        view._on_import_revision()

    rows = revisiones_repo.list_revision_rows(agente.id)
    assert len(rows) == 1
    assert rows[0].folio == "F-001"
    assert view.revision_table.rowCount() == 1


def test_import_revision_stores_and_displays_abogado_id(qapp, db, tmp_path):
    agente = _make_agente_abogado()
    abogado = users_repo.list_by_role(ROLE_ABOGADO)[0]
    path = tmp_path / "captura.mcdiep"
    _write_captura_file(path)

    view = RequerimientosRevisionView(agente)
    with patch(
        "app.ui.agente.requerimientos_revision_view.QFileDialog.getOpenFileName",
        return_value=(str(path), ""),
    ), patch("app.ui.agente.requerimientos_revision_view.QMessageBox.information"):
        view._on_import_revision()

    rows = revisiones_repo.list_revision_rows(agente.id)
    assert rows[0].abogado_id == abogado.id

    id_col = len(HEADERS_ABOGADO) + 1  # a la derecha de la columna Procede
    assert view.revision_table.item(0, id_col).text() == str(abogado.id)


def test_procede_combo_change_persists(qapp, db):
    agente = _make_agente_abogado()
    revision_import_id = _import_rows(agente.id, source_filename="x.xlsx", rows=[_row("F-001")])
    view = RequerimientosRevisionView(agente)
    view._load_import(revision_import_id)
    row_id = revisiones_repo.list_revision_rows(agente.id)[0].id
    combo = view.revision_table.cellWidget(0, len(HEADERS_ABOGADO))  # columna PROCEDE

    combo.setCurrentIndex(combo.findData("PROCEDE"))

    refreshed = revisiones_repo.list_revision_rows(agente.id)[0]
    assert refreshed.id == row_id
    assert refreshed.procede == "PROCEDE"


def test_export_revision_writes_file(qapp, db, tmp_path):
    from app.utils.paths import exports_dir

    agente = _make_agente_abogado()
    _import_rows(agente.id, source_filename="x.xlsx", rows=[_row("F-001")])
    view = RequerimientosRevisionView(agente)

    with patch("app.ui.agente.requerimientos_revision_view.QMessageBox.information"):
        view._on_export_revision()

    matches = list(exports_dir().glob("REVISION DEL *.xlsx"))
    assert len(matches) == 1


def test_export_revision_includes_despacho_and_observaciones(qapp, db, tmp_path):
    import openpyxl

    from app.excel_io.requerimientos_export import HEADERS_REVISION
    from app.utils.paths import exports_dir

    agente = _make_agente_abogado()
    row = _row("F-001")
    row["observaciones"] = "No se encontró a nadie en el domicilio."
    revision_import_id = revisiones_repo.create_revision_import(
        agente_id=agente.id, source_filename="x.xlsx", abogado_nombre="Abogado Uno", abogado_id=None,
    )
    revisiones_repo.add_revision_rows(
        agente_id=agente.id, revision_import_id=revision_import_id,
        source_filename="x.xlsx", abogado_nombre="Abogado Uno", abogado_id=None, rows=[row],
    )

    view = RequerimientosRevisionView(agente)
    with patch("app.ui.agente.requerimientos_revision_view.QMessageBox.information"):
        view._on_export_revision()

    output_path = next(exports_dir().glob("REVISION DEL *.xlsx"))
    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook.active
    header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header_row == HEADERS_REVISION

    despacho_col = HEADERS_REVISION.index("Despacho")
    observaciones_col = HEADERS_REVISION.index("Observaciones")
    data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert any(r[despacho_col] == "Abogado Uno" for r in data_rows)
    assert any(r[observaciones_col] == "No se encontró a nadie en el domicilio." for r in data_rows)


def test_simulate_mode_blocks_import_procede_and_export(qapp, db, tmp_path):
    agente = _make_agente_abogado()
    path = tmp_path / "captura.mcdiep"
    _write_captura_file(path)

    view = RequerimientosRevisionView(agente, simulate=True)

    with patch(
        "app.ui.agente.requerimientos_revision_view.QFileDialog.getOpenFileName"
    ) as mock_dialog, patch("app.ui.agente.requerimientos_revision_view.QMessageBox.information") as mock_info:
        view._on_import_revision()
    mock_dialog.assert_not_called()
    mock_info.assert_called_once()
    assert revisiones_repo.list_revision_rows(agente.id) == []

    # Import real preexistente (de otra sesión no-simulada) para probar que
    # el cambio de PROCEDE en modo simulación no persiste, aunque sí se
    # pueda navegar y abrir (de sólo lectura).
    revision_import_id = _import_rows(agente.id, source_filename="x.xlsx", rows=[_row("F-001")])
    view._refresh_available_list()
    view._load_import(revision_import_id)
    combo = view.revision_table.cellWidget(0, len(HEADERS_ABOGADO))
    combo.setCurrentIndex(combo.findData("PROCEDE"))
    assert revisiones_repo.list_revision_rows(agente.id)[0].procede is None

    with patch("app.ui.agente.requerimientos_revision_view.QMessageBox.information") as mock_info2:
        view._on_export_revision()
    mock_info2.assert_called_once()
    assert "Simulación" in mock_info2.call_args[0][1]


# --- Buscar fila por FOLIO/CTA PREDIAL/CONTRIBUYENTE -----------------------------------

def _rows_for_search():
    return [
        {
            "folio": "F-001", "cta_predial": "CP-001", "contribuyente": "Juan Pérez", "domicilio": "Calle 1",
            "fecha_citatorio": None, "recibe_citatorio": None, "recibe_citatorio_nombre": None,
            "fecha_notificacion": None, "quien_recibe": None, "quien_recibe_nombre": None,
        },
        {
            "folio": "F-002", "cta_predial": "CP-002", "contribuyente": "María López", "domicilio": "Calle 2",
            "fecha_citatorio": None, "recibe_citatorio": None, "recibe_citatorio_nombre": None,
            "fecha_notificacion": None, "quien_recibe": None, "quien_recibe_nombre": None,
        },
    ]


def test_search_revision_by_contribuyente_selects_matching_row(qapp, db):
    agente = _make_agente_abogado()
    revision_import_id = _import_rows(agente.id, source_filename="x.xlsx", rows=_rows_for_search())
    view = RequerimientosRevisionView(agente)
    view._load_import(revision_import_id)

    view.revision_search_field_combo.setCurrentIndex(
        view.revision_search_field_combo.findText("CONTRIBUYENTE")
    )
    view.revision_search_input.setText("lópez")
    view._on_search_revision()

    assert view.revision_table.currentRow() == 1


def test_search_revision_no_match_shows_information(qapp, db):
    agente = _make_agente_abogado()
    revision_import_id = _import_rows(agente.id, source_filename="x.xlsx", rows=_rows_for_search())
    view = RequerimientosRevisionView(agente)
    view._load_import(revision_import_id)
    view.revision_search_input.setText("no existe en ninguna fila")

    with patch("app.ui.agente.requerimientos_revision_view.QMessageBox.information") as mock_info:
        view._on_search_revision()

    mock_info.assert_called_once()


# --- Columnas redimensionables ----------------------------------------------------------

def test_revision_table_uses_interactive_resize_mode(qapp, db):
    from PySide6.QtWidgets import QHeaderView

    agente = _make_agente_abogado()
    view = RequerimientosRevisionView(agente)

    assert view.revision_table.horizontalHeader().sectionResizeMode(0) == QHeaderView.ResizeMode.Interactive
    assert view.revision_table.horizontalHeader().stretchLastSection() is True


# --- Nombre del archivo en revisión (label + señal para el título de pestaña) ----------

def test_import_revision_updates_filename_label(qapp, db, tmp_path):
    from app.ui.agente.requerimientos_revision_view import SIN_ARCHIVO_TEXT

    agente = _make_agente_abogado()
    path = tmp_path / "captura.mcdiep"
    _write_captura_file(path)

    view = RequerimientosRevisionView(agente)
    assert view.filename_label.text() == SIN_ARCHIVO_TEXT

    with patch(
        "app.ui.agente.requerimientos_revision_view.QFileDialog.getOpenFileName",
        return_value=(str(path), ""),
    ), patch("app.ui.agente.requerimientos_revision_view.QMessageBox.information"):
        view._on_import_revision()

    assert view.filename_label.text() == f"Archivo en revisión: {path.name}"


def test_import_revision_emits_archivo_cambiado_signal(qapp, db, tmp_path):
    agente = _make_agente_abogado()
    path = tmp_path / "captura.mcdiep"
    _write_captura_file(path)

    view = RequerimientosRevisionView(agente)
    received = []
    view.archivo_cambiado.connect(received.append)

    with patch(
        "app.ui.agente.requerimientos_revision_view.QFileDialog.getOpenFileName",
        return_value=(str(path), ""),
    ), patch("app.ui.agente.requerimientos_revision_view.QMessageBox.information"):
        view._on_import_revision()

    assert received == [path.name]


# --- Corrección de la concatenación: cada import se ve por separado --------------------

def test_importing_second_file_does_not_concatenate_first_files_rows(qapp, db, tmp_path):
    agente = _make_agente_abogado()
    path1 = tmp_path / "captura1.mcdiep"
    _write_captura_file(path1, folio="F-001")
    path2 = tmp_path / "captura2.mcdiep"
    _write_captura_file(path2, folio="F-002")

    view = RequerimientosRevisionView(agente)

    with patch(
        "app.ui.agente.requerimientos_revision_view.QFileDialog.getOpenFileName",
        return_value=(str(path1), ""),
    ), patch("app.ui.agente.requerimientos_revision_view.QMessageBox.information"):
        view._on_import_revision()

    assert view.revision_table.rowCount() == 1
    assert view.revision_table.item(0, 0).text() == "F-001"

    with patch(
        "app.ui.agente.requerimientos_revision_view.QFileDialog.getOpenFileName",
        return_value=(str(path2), ""),
    ), patch("app.ui.agente.requerimientos_revision_view.QMessageBox.information"):
        view._on_import_revision()

    # Sólo el archivo recién importado, NO ambos concatenados.
    assert view.revision_table.rowCount() == 1
    assert view.revision_table.item(0, 0).text() == "F-002"

    # Pero ambos siguen existiendo por separado en la base.
    assert len(revisiones_repo.list_revision_rows(agente.id)) == 2


# --- Pantalla previa: estado (Pendiente/Revisado) + lista + Abrir/Limpiar --------------

def test_available_list_filters_by_estado_and_moves_when_fully_reviewed(qapp, db):
    from app.ui.agente.requerimientos_revision_view import DOC_STATE_PENDIENTE, DOC_STATE_REVISADO

    agente = _make_agente_abogado()
    import_id = _import_rows(agente.id, source_filename="lote1.xlsx", rows=[_row("F1")])

    view = RequerimientosRevisionView(agente)

    # Por defecto arranca en "Pendiente".
    assert view.available_list.count() == 1
    assert view.available_list.item(0).data(1000) == import_id

    view.estado_combo.setCurrentIndex(view.estado_combo.findData(DOC_STATE_REVISADO))
    assert view.available_list.count() == 0

    row_id = revisiones_repo.list_revision_rows_for_import(import_id)[0].id
    view.estado_combo.setCurrentIndex(view.estado_combo.findData(DOC_STATE_PENDIENTE))
    view._load_import(import_id)
    combo = view.revision_table.cellWidget(0, len(HEADERS_ABOGADO))
    combo.setCurrentIndex(combo.findData("PROCEDE"))

    # Al quedar completamente revisado, desaparece de "Pendiente"...
    assert view.available_list.count() == 0
    # ...y aparece en "Revisado".
    view.estado_combo.setCurrentIndex(view.estado_combo.findData(DOC_STATE_REVISADO))
    assert view.available_list.count() == 1
    assert view.available_list.item(0).data(1000) == import_id
    del row_id  # sólo usado para disparar el update arriba


def test_changing_estado_directly_clears_previous_available_list(qapp, db):
    from app.ui.agente.requerimientos_revision_view import DOC_STATE_REVISADO

    agente = _make_agente_abogado()
    _import_rows(agente.id, source_filename="lote1.xlsx", rows=[_row("F1")])  # PENDIENTE

    view = RequerimientosRevisionView(agente)
    assert view.available_list.count() == 1

    # No hay ningún import REVISADO todavía: cambiar el estado directamente
    # debe vaciar la lista mostrada.
    view.estado_combo.setCurrentIndex(view.estado_combo.findData(DOC_STATE_REVISADO))
    assert view.available_list.count() == 0


def test_open_selected_import_loads_table(qapp, db):
    agente = _make_agente_abogado()
    _import_rows(agente.id, source_filename="lote1.xlsx", rows=[_row("F1"), _row("F2")])

    view = RequerimientosRevisionView(agente)
    assert view.available_list.count() == 1
    view.available_list.setCurrentRow(0)
    view._on_open_selected_import()

    assert view.revision_table.rowCount() == 2


def test_open_with_nothing_selected_shows_information(qapp, db):
    agente = _make_agente_abogado()
    _import_rows(agente.id, source_filename="lote1.xlsx", rows=[_row("F1")])

    view = RequerimientosRevisionView(agente)
    view.available_list.setCurrentRow(-1)

    with patch("app.ui.agente.requerimientos_revision_view.QMessageBox.information") as mock_info:
        view._on_open_selected_import()

    mock_info.assert_called_once()


def test_clear_resets_list_and_open_table(qapp, db, tmp_path):
    from app.ui.agente.requerimientos_revision_view import SIN_ARCHIVO_TEXT

    agente = _make_agente_abogado()
    path = tmp_path / "captura.mcdiep"
    _write_captura_file(path)

    view = RequerimientosRevisionView(agente)
    with patch(
        "app.ui.agente.requerimientos_revision_view.QFileDialog.getOpenFileName",
        return_value=(str(path), ""),
    ), patch("app.ui.agente.requerimientos_revision_view.QMessageBox.information"):
        view._on_import_revision()
    assert view.revision_table.rowCount() == 1

    view._on_clear()

    assert view.available_list.count() == 0
    assert view._current_import_id is None
    assert view.revision_table.rowCount() == 0
    assert view.filename_label.text() == SIN_ARCHIVO_TEXT
